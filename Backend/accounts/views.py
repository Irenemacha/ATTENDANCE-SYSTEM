import random
import string
from datetime import timedelta
from io import BytesIO
from django.core.mail import send_mail
from students.models import Notification
from accounts.permissions import IsStudent

from django.contrib.auth import authenticate, get_user_model
from django.contrib.auth.models import Group, Permission
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenRefreshView
from attendance.models import Attendance, AttendanceSession, MovementLog

from accounts.models import OTP, UserDevice, UserSessionState
from accounts.permissions import IsAdminOrStaff
from accounts.serializers import (
    GroupSerializer,
    PermissionSerializer,
    UserListSerializer,
    UserMeSerializer,
    UserWriteSerializer,
)
from accounts.services import advance_user_state
from courses.models import Course
from students.models import Student

User = get_user_model()


@api_view(["GET"])
@permission_classes([AllowAny])
def api_root(request):
    return Response({
        "auth": {
            "login": "/api/auth/login/",
            "refresh": "/api/auth/refresh/",
            "me": "/api/auth/me/",
        },
        "user_management": {
            "users": "/api/user-management/users/",
            "groups": "/api/user-management/groups/",
            "permissions": "/api/user-management/permissions/",
            "import_preview": "/api/user-management/users/import/preview/",
            "import_commit": "/api/user-management/users/import/commit/",
        },
        "attendance": {
            "check_in": "/api/attendance/check-in/",
            "check_out": "/api/attendance/check-out/",
            "mark": "/api/attendance/mark/",
            "start_session": "/api/attendance/start-session/",
            "end_session": "/api/attendance/end-session/",
            "report": "/api/attendance/report/",
        },
        "courses": {
            "list_create": "/api/courses/",
            "assign_student": "/api/courses/assign-student/",
            "assign_lecturer": "/api/courses/assign-lecturer/",
        },
        "students": {
            "dashboard": "/api/students/dashboard/",
        },
    })


@api_view(["POST"])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")
    if not username or not password:
        return Response(
            {"detail": "Username and password are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = authenticate(request, username=username, password=password)
    if not user:
        return Response(
            {"detail": "Invalid username or password."},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    refresh = RefreshToken.for_user(user)
    return Response({
        "access": str(refresh.access_token),
        "refresh": str(refresh),
        "user": UserMeSerializer(user).data,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me(request):
    return Response(UserMeSerializer(request.user).data)


@api_view(["POST"])
@permission_classes([AllowAny])
def device_login(request):
    username = request.data.get("username")
    user = User.objects.get(username=username)
    password = request.data.get("password")
    device_id = request.data.get("device_id")

    user = User.objects.filter(username=username).first()

    print("========== LOGIN DEBUG ==========")
    print("USERNAME RECEIVED:", username)
    print("PASSWORD RECEIVED:", password)
    print("USER FOUND:", user)

    if user:
        print("PASSWORD CHECK:", user.check_password(password))

    if not user or not user.check_password(password):
        print("LOGIN FAILED HERE")
        return Response({"error": "Invalid credentials"}, status=400)

    print("LOGIN PASSED PASSWORD CHECK")

    if not device_id:
        return Response({"error": "Device ID required"}, status=400)
    
    existing_device = UserDevice.objects.filter(
    device_id=device_id
).first()

    if existing_device and existing_device.user != user:
       return Response(
        {
            "success": False,
            "message": "This device is already registered to another user."
        },
        status=400
        )

    device = UserDevice.objects.filter(
        user=user,
        device_id=device_id,
        is_verified=True,
    ).first()

    # ==========================
    # DEVICE ALREADY VERIFIED
    # ==========================
    if device:
        advance_user_state(user, "device_success")
        refresh = RefreshToken.for_user(user)

        return Response({
            "message": "Login successful (device already verified)",
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": UserMeSerializer(user).data,
            "device_required": False,
        })

    # ==========================
    # CREATE DEVICE ENTRY
    # ==========================
    UserDevice.objects.get_or_create(user=user, device_id=device_id)

    # ==========================
    # OTP GENERATION
    otp_code = str(random.randint(100000, 999999))

    otp_obj = OTP.objects.create(
        user=user,
        code=otp_code,
        purpose="device_verification",
        expires_at=timezone.now() + timedelta(minutes=5),
   )

    print("DEBUG OTP:", otp_obj.code)
    print("EMAIL:", user.email)

    print("ABOUT TO SEND OTP EMAIL")

    send_mail(
        subject="Your OTP Code",
        message=f"Your OTP is {otp_code}. It expires in 5 minutes.",
        from_email="irenemagige548@gmail.com",
        recipient_list=[user.email],
        fail_silently=False,
    )

    print("OTP EMAIL SENT")
    # ==========================
    # DEBUG PRINTS (IMPORTANT)
    # ==========================
    print("====================================")
    print("DEVICE LOGIN HIT")
    print("OTP GENERATED:", otp_obj.code)
    print("DEVICE ID:", device_id)
    print("====================================")

    advance_user_state(user, "otp_sent")

    # ==========================
    # RESPONSE (FIXED)
    # ==========================
    return Response({
        "success": True,
        "message": "OTP generated and required for device verification",
        "device_required": True,
        "demoOtp": otp_obj.code,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generate_otp(request):
    print("ABOUT TO GENERATE OTP")
    otp_obj = OTP.objects.create(
        user=request.user,
        code=str(random.randint(100000, 999999)),
        purpose="device_verification",
        expires_at=timezone.now() + timedelta(minutes=5),
    )
    print("OTP GENERATED (device_login):", otp_obj.code)
    return Response({"message": "OTP generated", "device_required": True, "demoOtp": otp_obj.code})




@api_view(["POST"])
@permission_classes([AllowAny])
def verify_otp(request):
    username = request.data.get("username")
    otp_input = request.data.get("otp")
    device_id = request.data.get("device_id")

    if not username:
        return Response({"error": "Username is required"}, status=400)

    if not otp_input:
        return Response({"error": "OTP is required"}, status=400)

    user = User.objects.filter(username=username).first()
    print("USER FOUND:", user is not None)

    if user:
        print("USER EMAIL:", user.email)
    if not user:
        return Response({"error": "User not found"}, status=400)

    # ALWAYS get latest OTP for this user ONLY
    otp_obj = OTP.objects.filter(user=user).order_by("-id").first()

    if not otp_obj:
        return Response({"error": "No OTP found"}, status=400)

    if otp_obj.is_expired():
        return Response({"error": "OTP expired"}, status=400)

    if otp_obj.is_used:
        return Response({"error": "OTP already used"}, status=400)

    if str(otp_obj.code) != str(otp_input):
        return Response({"error": "Invalid OTP"}, status=400)

    otp_obj.is_used = True
    otp_obj.save(update_fields=["is_used"])

    state, _ = UserSessionState.objects.get_or_create(user=user)
    state.otp_verified = True
    state.current_state = "ATTENDANCE_GRANTED"
    state.save()

    if device_id:
        UserDevice.objects.update_or_create(
            user=user,
            device_id=device_id,
            defaults={"is_verified": True},
        )

    advance_user_state(user, "otp_success")

    refresh = RefreshToken.for_user(user)

    return Response({
        "message": "OTP verified successfully",
        "access": str(refresh.access_token),
        "refresh": str(refresh),
    })
    
@api_view(["POST"])
@permission_classes([AllowAny])
def verify_device_otp(request):
    username = request.data.get("username")
    otp_input = request.data.get("otp")
    device_id = request.data.get("device_id")
    user = User.objects.filter(username=username).first()
    if not user:
        return Response({"error": "User not found"}, status=400)

    otp_obj = OTP.objects.filter(user=user).order_by("-id").first()
    if not otp_obj:
        return Response({"error": "No OTP found"}, status=400)
    if str(otp_obj.code) != str(otp_input):
        return Response({"error": "Invalid OTP"}, status=400)
    if otp_obj.is_expired():
        otp_obj.delete()
        return Response({"error": "OTP expired"}, status=400)

    otp_obj.is_used = True
    otp_obj.save(update_fields=["is_used"])
    if device_id:
        UserDevice.objects.update_or_create(
            user=user,
            device_id=device_id,
            defaults={"is_verified": True},
        )
    advance_user_state(user, "device_success")
    refresh = RefreshToken.for_user(user)
    return Response({
        "message": "OTP verified successfully",
        "access": str(refresh.access_token),
        "refresh": str(refresh),
        "user": UserMeSerializer(user).data,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def fingerprint_verify(request):

    success = request.data.get("success", True)

    state, _ = UserSessionState.objects.get_or_create(
        user=request.user
    )

    if success in [True, "true", "True", 1, "1"]:

        advance_user_state(
            request.user,
            "fingerprint_success"
        )

        return Response({
            "message": "Fingerprint verified",
            "state": "ATTENDANCE_GRANTED"
        })


    state = advance_user_state(
        request.user,
        "fingerprint_fail"
    )

    return Response({
        "message": "Fingerprint failed",
        "state": state.current_state
    }, status=403)


@api_view(["GET"])
@permission_classes([AllowAny])
def test(request):
    return Response({"message": "working"})


class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.prefetch_related("groups").order_by("id")
    permission_classes = [IsAdminOrStaff]
    pagination_class = StandardPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["username", "email", "first_name", "last_name", "groups__name"]

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return UserWriteSerializer
        return UserListSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        group_name = self.request.query_params.get("group")
        if group_name:
            queryset = queryset.filter(groups__name__iexact=group_name)
        return queryset.distinct()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserListSerializer(user).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        user = self.get_object()
        serializer = self.get_serializer(user, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserListSerializer(user).data)

    @action(detail=True, methods=["patch"])
    def groups(self, request, pk=None):
        user = self.get_object()
        group_names = request.data.get("groups", [])
        if not isinstance(group_names, list):
            return Response({"detail": "groups must be a list."}, status=400)
        groups = list(Group.objects.filter(name__in=group_names))
        found_names = {group.name for group in groups}
        missing = [name for name in group_names if name not in found_names]
        if missing:
            return Response({"detail": f"Unknown groups: {', '.join(missing)}"}, status=400)
        user.groups.set(groups)
        return Response(UserListSerializer(user).data)

    @action(detail=False, methods=["post"], url_path="import/preview")
    def import_preview(self, request):
        rows, errors = parse_import_rows(request)
        return Response({"rows": rows, "errors": errors, "valid": not errors})

    @action(detail=False, methods=["post"], url_path="import/commit")
    def import_commit(self, request):
        update_existing = str(request.data.get("update_existing", "false")).lower() == "true"
        rows, parse_errors = parse_import_rows(request)
        result = {
            "created_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "errors": list(parse_errors),
            "created_users": [],
        }
        if parse_errors and not request.data.get("allow_partial"):
            result["skipped_count"] = len(rows)
            return Response(result, status=400)

        with transaction.atomic():
            for row in rows:
                if row["errors"]:
                    result["errors"].append({"row": row["row"], "errors": row["errors"]})
                    result["skipped_count"] += 1
                    continue
                created_password = None
                user = User.objects.filter(username=row["username"]).first()
                if user and not update_existing:
                    result["errors"].append({
                        "row": row["row"],
                        "errors": ["Duplicate username; skipped."],
                    })
                    result["skipped_count"] += 1
                    continue

                password = row.get("password") or generate_default_password()
                if not row.get("password"):
                    created_password = password

                if user:
                    user.email = row.get("email", "")
                    user.first_name = row.get("first_name", "")
                    user.last_name = row.get("last_name", "")
                    if row.get("password"):
                        user.set_password(password)
                    user.save()
                    result["updated_count"] += 1
                else:
                    user = User.objects.create_user(
                        username=row["username"],
                        password=password,
                        email=row.get("email", ""),
                        first_name=row.get("first_name", ""),
                        last_name=row.get("last_name", ""),
                    )
                    result["created_count"] += 1

                assign_groups(user, row["groups"])
                if any(name.lower() == "student" for name in row["groups"]):
                    upsert_student_profile(user, row)

                payload = UserListSerializer(user).data
                if created_password:
                    payload["generated_password"] = created_password
                result["created_users"].append(payload)

        return Response(result)


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.prefetch_related("permissions").order_by("name")
    serializer_class = GroupSerializer
    permission_classes = [IsAdminOrStaff]
    pagination_class = None

    @action(detail=True, methods=["patch"])
    def permissions(self, request, pk=None):
        group = self.get_object()
        permission_ids = request.data.get("permissions", [])
        if not isinstance(permission_ids, list):
            return Response({"detail": "permissions must be a list."}, status=400)
        permissions = Permission.objects.filter(id__in=permission_ids)
        group.permissions.set(permissions)
        return Response(GroupSerializer(group).data)


class PermissionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Permission.objects.select_related("content_type").order_by(
        "content_type__app_label",
        "content_type__model",
        "codename",
    )
    serializer_class = PermissionSerializer
    permission_classes = [IsAdminOrStaff]
    pagination_class = None


def parse_import_rows(request):
    if "file" not in request.FILES:
        return [], [{"row": None, "errors": ["Excel file is required."]}]

    workbook = load_workbook(BytesIO(request.FILES["file"].read()), data_only=True)
    sheet = workbook.active
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    expected = {
        "username",
        "password",
        "email",
        "first_name",
        "last_name",
        "groups",
        "reg_number",
        "phone_number",
        "course_code",
        "year_of_study",
    }
    missing_required = [name for name in ["username"] if name not in headers]
    global_errors = []
    if missing_required:
        global_errors.append({"row": None, "errors": [f"Missing columns: {', '.join(missing_required)}"]})

    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(headers, values))
        row = normalize_import_row(index, raw)
        unknown_columns = [header for header in headers if header and header not in expected]
        if unknown_columns:
            row["errors"].append(f"Unsupported columns: {', '.join(unknown_columns)}")
        rows.append(row)
    return rows, global_errors


def normalize_import_row(index, raw):
    groups = [
        value.strip()
        for value in str(raw.get("groups") or "").replace(";", ",").split(",")
        if value and value.strip()
    ]
    row = {
        "row": index,
        "username": str(raw.get("username") or "").strip(),
        "password": str(raw.get("password") or "").strip(),
        "email": str(raw.get("email") or "").strip(),
        "first_name": str(raw.get("first_name") or "").strip(),
        "last_name": str(raw.get("last_name") or "").strip(),
        "groups": groups,
        "reg_number": str(raw.get("reg_number") or "").strip(),
        "phone_number": str(raw.get("phone_number") or "").strip(),
        "course_code": str(raw.get("course_code") or "").strip(),
        "year_of_study": raw.get("year_of_study") or None,
        "errors": [],
    }
    if not row["username"]:
        row["errors"].append("username is required.")
    missing_groups = [
        name for name in groups if not Group.objects.filter(name__iexact=name).exists()
    ]
    if missing_groups:
        row["errors"].append(f"Unknown groups: {', '.join(missing_groups)}")
    if any(name.lower() == "student" for name in groups):
        for field in ["reg_number", "course_code", "year_of_study"]:
            if not row[field]:
                row["errors"].append(f"{field} is required for Student users.")
        if row["course_code"] and not Course.objects.filter(code__iexact=row["course_code"]).exists():
            row["errors"].append("course_code does not match an existing course.")
    return row


def assign_groups(user, group_names):
    groups = Group.objects.none()
    for name in group_names:
        groups = groups | Group.objects.filter(name__iexact=name)
    user.groups.set(groups)


def upsert_student_profile(user, row):
    course = Course.objects.get(code__iexact=row["course_code"])
    full_name = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip() or user.username
    Student.objects.update_or_create(
        user=user,
        defaults={
            "reg_number": row["reg_number"],
            "full_name": full_name,
            "email": row.get("email") or user.email,
            "phone_number": row.get("phone_number") or "",
            "course": course,
            "year_of_study": int(row["year_of_study"]),
        },
    )


def generate_default_password(length=12):
    alphabet = string.ascii_letters + string.digits
    return "".join(random.choice(alphabet) for _ in range(length))


refresh = TokenRefreshView.as_view()

@api_view(["GET"])
@permission_classes([IsStudent])
def notifications(request):

    student = request.user.student

    notifications = Notification.objects.filter(
        student=student
    ).order_by("-created_at")

    return Response([
        {
            "title": n.title,
            "message": n.message,
            "read": n.is_read,
            "date": n.created_at
        }
        for n in notifications
    ])